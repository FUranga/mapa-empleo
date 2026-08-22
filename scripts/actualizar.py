"""
actualizar.py
Corre en GitHub Actions. Verifica si hay datos nuevos en SRT y OEDE,
los baja, regenera los JSONs y los guarda en el repo.
"""

import requests
import openpyxl
import json
import os
import sys
from datetime import datetime
from io import BytesIO

# ── URLs de las fuentes ───────────────────────────────────────────
URL_SRT_JURISDICCION = (
    'https://www.srt.gob.ar/estadisticas/series/co/up/'
    'Serie_historica_Segun_Jurisdiccion - Ubicacion Persona Trabajadora - UP.xlsx'
)
URL_SRT_SECTOR = (
    'https://www.srt.gob.ar/estadisticas/series/co/up/'
    'Serie_historica_Segun_Sector_de_actividad_economica_CIIUrev4 - UP.xlsx'
)

# OEDE publica en una URL fija que sobreescribe con cada actualización
# Verificar esta URL con cada publicación nueva
URL_OEDE = (
    'https://www.trabajo.gob.ar/downloads/estadisticas/'
    'observatorio/series/cuadros_empleo_privado.xlsx'
)

LOG_PATH    = 'log.json'
EMP_PATH    = 'data.json'
EMP_PATH_E  = 'empresas.json'

# ── Logging ───────────────────────────────────────────────────────
def load_log():
    if os.path.exists(LOG_PATH):
        with open(LOG_PATH) as f:
            return json.load(f)
    return {}

def save_log(log):
    with open(LOG_PATH, 'w') as f:
        json.dump(log, f, ensure_ascii=False, indent=2)

def get_last_modified(url):
    """HEAD request para ver fecha de modificación del archivo remoto."""
    try:
        r = requests.head(url, timeout=20, allow_redirects=True)
        return r.headers.get('Last-Modified') or r.headers.get('ETag') or ''
    except Exception as e:
        print(f'  ⚠ No se pudo verificar {url}: {e}')
        return None

def download(url):
    """Descarga un archivo y devuelve sus bytes."""
    print(f'  Descargando: {url[:80]}...')
    r = requests.get(url, timeout=120)
    r.raise_for_status()
    return r.content

# ── Generador de empresas.json ────────────────────────────────────
def build_empresas(bytes_juris, bytes_sector):
    from datetime import datetime as dt

    def pct(base, curr):
        if base and base > 0:
            return round((curr - base) / base * 100, 1)
        return None

    # Jurisdicción
    wb_j = openpyxl.load_workbook(BytesIO(bytes_juris), data_only=True)
    ws_j = wb_j['Cuadro 6.2']
    rows_j = list(ws_j.iter_rows(values_only=True))
    header = rows_j[4]
    periodos = [v.strftime('%Y-%m') for v in header[1:] if isinstance(v, dt)]

    SKIP = {'Sin datos', None,
        'Parte empleadora afiliada de unidades productivas con personas trabajadoras declaradas *',
        'Parte empleadora afiliada y aportante de casas particulares con personas trabajadoras declaradas **',
        'Parte empleadora afiliada de casas particulares con personas trabajadoras declaradas ***'}

    prov_data = {}
    for row in rows_j[5:]:
        nombre = row[0]
        if not isinstance(nombre, str) or nombre in SKIP: continue
        if nombre.startswith('Debido') or nombre.startswith('*') or nombre.startswith('Fuente'): continue
        vals = row[1:len(periodos)+1]
        if not any(isinstance(v, (int, float)) for v in vals): continue
        prov_data[nombre] = {periodos[i]: int(v) for i, v in enumerate(vals) if isinstance(v, (int, float))}

    # Sectores
    wb_s = openpyxl.load_workbook(BytesIO(bytes_sector), data_only=True)
    ws_s = wb_s['Cuadro 2.2']
    rows_s = list(ws_s.iter_rows(values_only=True))
    header_s = rows_s[4]
    periodos_s = [v.strftime('%Y-%m') for v in header_s[1:] if isinstance(v, dt)]

    SKIP_SEC = {
        'Parte empleadora afiliada de unidades productivas con personas trabajadoras declaradas  (1) *',
        'Parte empleadora afiliada y aportante de casas particulares con personas trabajadoras declaradas (2) **',
        'Parte empleadora afiliada de casas particulares con personas trabajadoras declaradas (3)***',
        'Parte empleadora afiliada con personas trabajadoras declaradas = (1) + (3)',
        'Total parte empleadora afiliada del sistema****', 'Sin datos',
    }

    SECTOR_CORTO = {
        'Agricultura, ganaderia, caza, silvicultura y pesca': 'Agro y pesca',
        'Explotacion de minas y canteras': 'Minería',
        'Industria manufacturera': 'Industria',
        'Suministro de electricidad, gas, vapor y aire acondicionado': 'Energía eléctrica',
        'Suministro de agua, cloacas, gestion de residuos y recuperacion de materiales y saneamiento publico': 'Agua y saneamiento',
        'Construccion': 'Construcción',
        'Comercio al por mayor y al por menor; reparacion de vehiculos automotores y motocicletas': 'Comercio',
        'Servicio de transporte y almacenamiento': 'Transporte',
        'Servicios de alojamiento y servicios de comida': 'Alojamiento y gastronomía',
        'Informacion y comunicaciones': 'Info. y comunicaciones',
        'Intermediacion financiera y servicios de seguros': 'Finanzas y seguros',
        'Servicios inmobiliarios': 'Inmobiliario',
        'Servicios profesionales, cientificos y tecnicos': 'Servicios profesionales',
        'Actividades administrativas y servicios de apoyo': 'Serv. administrativos',
        # 'Administracion publica, defensa y seguridad social obligatoria': 'Administración pública',  # excluido — empleador público
        'Enseñanza': 'Enseñanza',
        'Salud humana y servicios sociales': 'Salud',
        'Servicios artisticos, culturales, deportivos y de esparcimiento': 'Arte y esparcimiento',
        'Servicios de asociaciones y servicios personales': 'Asoc. y serv. personales',
        # 'Servicios de organizaciones y organos extraterritoriales': 'Org. extraterritoriales',  # excluido — organizaciones extraterritoriales
    }

    sec_data = {}
    total_nac_s = {}
    TOTAL_ROW = 26  # fila exacta del total nacional en Cuadro 2.2
    for i, row in enumerate(rows_s):
        nombre = row[0]
        if not isinstance(nombre, str): continue
        if i == TOTAL_ROW:
            for j, v in enumerate(row[1:len(periodos_s)+1]):
                if isinstance(v, (int, float)):
                    total_nac_s[periodos_s[j]] = int(v)
            continue
        if not (5 <= i <= 24): continue
        corto = SECTOR_CORTO.get(nombre, nombre)
        vals = row[1:len(periodos_s)+1]
        if any(isinstance(v, (int, float)) for v in vals):
            sec_data[corto] = {periodos_s[j]: int(v) for j, v in enumerate(vals) if isinstance(v, (int, float))}

    ULTIMO = periodos[-1]
    PRESIDENCIAS = {
        'Fernández': {'inicio': '2019-11', 'fin': '2023-11'},
        'Milei':     {'inicio': '2023-11', 'fin': ULTIMO},
    }

    serie_nac = [{'t': t, 'v': total_nac_s.get(t)} for t in periodos_s]
    sectores_list = [
        {'sector': nombre, 'serie': [{'t': t, 'v': serie.get(t)} for t in periodos_s]}
        for nombre, serie in sec_data.items()
    ]
    provincias_obj = {
        nombre: {'serie': [{'t': t, 'v': serie.get(t)} for t in periodos]}
        for nombre, serie in prov_data.items()
    }

    return {
        'meta': {
            'ultimo':        periodos_s[-1],
            'periodos':      periodos_s,
            'periodos_prov': periodos,
            'presidencias':  PRESIDENCIAS,
            'fuente': 'SRT — Superintendencia de Riesgos del Trabajo',
            'actualizado': datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'),
        },
        'pais': {'serie': serie_nac, 'sectores': sectores_list},
        'provincias': provincias_obj,
    }

# ── Main ──────────────────────────────────────────────────────────
def main():
    log = load_log()
    today = datetime.utcnow().strftime('%Y-%m-%d')
    updated = False

    print('=== Verificando SRT (empresas) ===')
    mod_juris  = get_last_modified(URL_SRT_JURISDICCION)
    mod_sector = get_last_modified(URL_SRT_SECTOR)
    srt_sig    = f'{mod_juris}|{mod_sector}'

    if srt_sig and srt_sig != log.get('srt_signature'):
        print('  ✓ Hay datos nuevos en SRT — descargando...')
        try:
            bytes_juris  = download(URL_SRT_JURISDICCION)
            bytes_sector = download(URL_SRT_SECTOR)
            print('  Construyendo empresas.json...')
            empresas = build_empresas(bytes_juris, bytes_sector)
            with open(EMP_PATH_E, 'w', encoding='utf-8') as f:
                json.dump(empresas, f, ensure_ascii=False, separators=(',', ':'))
            print(f'  ✓ empresas.json actualizado — último período: {empresas["meta"]["ultimo"]}')
            log['srt_signature'] = srt_sig
            log['srt_last_update'] = today
            updated = True
        except Exception as e:
            print(f'  ✗ Error en SRT: {e}')
    else:
        print(f'  Sin cambios en SRT (última actualización: {log.get("srt_last_update", "nunca")})')

    print('\n=== Verificando OEDE (empleo) ===')
    try:
        # Scrapeamos la página para encontrar las URLs actuales (cambian con cada publicación)
        PAGE_SIPA = 'https://www.argentina.gob.ar/trabajo/estadisticas/situacion-y-evolucion-del-trabajo-registrado'
        PAGE_DEPT = 'https://www.argentina.gob.ar/trabajo/estadisticas/oede-estadisticas-provinciales'
        import re as _re

        def find_url(page_url, pattern):
            r = requests.get(page_url, timeout=30)
            matches = _re.findall(pattern, r.text)
            if matches:
                path = matches[0]
                return 'https://www.argentina.gob.ar' + path if path.startswith('/') else path
            return None

        # SIPA mensual: trabajoregistrado_AAMM_estadisticas.xlsx
        url_nac  = find_url(PAGE_SIPA, r'(/sites/default/files/trabajoregistrado_\d+_estadisticas\.xlsx)')
        url_dept = 'https://raw.githubusercontent.com/FUranga/mapa-empleo/main/departamento_series_empleo_y_salarios_mensual_sector_1.csv'

        print(f'  URL SIPA mensual:  {url_nac}')
        print(f'  URL departamental: {url_dept}')

        # Alerta si no se encuentra la URL del SIPA
        if not url_nac:
            print('  ✗ No se encontró la URL del SIPA — el patrón puede haber cambiado')
            log['sipa_url_missing'] = True
            log['sipa_url_missing_date'] = today
        else:
            log.pop('sipa_url_missing', None)

        oede_sig = url_nac or ''

        # Verificar CSV departamental (URL estable)
        url_dept_oede = 'https://www.argentina.gob.ar/sites/default/files/departamento_series_empleo_y_salarios_mensual_sector_1.csv'
        mod_dept = get_last_modified(url_dept_oede)
        dept_sig = mod_dept or ''
        if dept_sig and dept_sig != log.get('dept_signature'):
            print(f'  📬 CSV departamental actualizado — avisando por mail')
            log['dept_signature'] = dept_sig
            log['dept_last_detected'] = today
            log['dept_needs_manual_update'] = True
            updated_dept = True
        else:
            updated_dept = False
            print(f'  Sin cambios en departamental (última detección: {log.get("dept_last_detected", "nunca")})')

        # Verificar XLSX del OEDE en la página del gobierno
        PAGE_PROV = 'https://www.argentina.gob.ar/trabajo/estadisticas/oede-estadisticas-provinciales'
        try:
            import re as _re
            r_prov = requests.get(PAGE_PROV, timeout=30)
            pat_dept = r"(/sites/default/files/departamento_serie_empleo_remuneraciones[^ ]+[.]xlsx)"
            pat_trim = r"(/sites/default/files/provinciales_serie_empleo_trimestral_2dig[^ ]+[.]xlsx)"
            found_dept = _re.search(pat_dept, r_prov.text)
            found_trim = _re.search(pat_trim, r_prov.text)
            url_xlsx_dept_web = found_dept.group(1) if found_dept else None
            url_xlsx_trim_web = found_trim.group(1) if found_trim else None

            # XLSX departamental
            if not url_xlsx_dept_web:
                print('  ✗ XLSX departamental no encontrado en página OEDE — URL puede haber cambiado')
                log['xlsx_dept_web_missing'] = True
                log['xlsx_dept_web_missing_date'] = today
            elif url_xlsx_dept_web != log.get('xlsx_dept_web_url'):
                print(f'  📬 XLSX departamental nuevo en web: {url_xlsx_dept_web}')
                log['xlsx_dept_web_url'] = url_xlsx_dept_web
                log['xlsx_dept_web_detected'] = today
                log['xlsx_dept_needs_manual_update'] = True
                log.pop('xlsx_dept_web_missing', None)
            else:
                print(f'  Sin cambios en XLSX departamental web')
                log.pop('xlsx_dept_web_missing', None)

            # XLSX trimestral provincial
            if not url_xlsx_trim_web:
                print('  ✗ XLSX trimestral provincial no encontrado en página OEDE — URL puede haber cambiado')
                log['xlsx_trim_web_missing'] = True
                log['xlsx_trim_web_missing_date'] = today
            elif url_xlsx_trim_web != log.get('xlsx_trim_web_url'):
                print(f'  📬 XLSX trimestral provincial nuevo en web: {url_xlsx_trim_web}')
                log['xlsx_trim_web_url'] = url_xlsx_trim_web
                log['xlsx_trim_web_detected'] = today
                log['xlsx_trim_needs_manual_update'] = True
                log.pop('xlsx_trim_web_missing', None)
            else:
                print(f'  Sin cambios en XLSX trimestral provincial web')
                log.pop('xlsx_trim_web_missing', None)

        except Exception as e:
            print(f'  ⚠ No se pudo verificar página OEDE: {e}')

        if oede_sig != log.get('oede_signature'):
            print('  ✓ Hay datos nuevos en OEDE — descargando...')
            bytes_nac  = download(url_nac)  if url_nac  else None
            bytes_dept = download(url_dept)
            # XLSX de totales departamentales (opcional, mejora precisión)
            url_xlsx = 'https://raw.githubusercontent.com/FUranga/mapa-empleo/main/departamento_serie_empleo_remuneraciones_3.xlsx'
            try:
                bytes_xlsx = download(url_xlsx)
            except:
                bytes_xlsx = None
                print('  ✗ XLSX departamental no encontrado en repo')
                log['xlsx_dept_missing'] = True
                log['xlsx_dept_missing_date'] = today

            # XLSX trimestral provincial (sectores por provincia)
            url_prov_trim = 'https://raw.githubusercontent.com/FUranga/mapa-empleo/main/provinciales_serie_empleo_trimestral_2dig_6.xlsx'
            try:
                bytes_prov_trim = download(url_prov_trim)
                log.pop('xlsx_prov_trim_missing', None)
            except:
                bytes_prov_trim = None
                print('  ✗ Trimestral provincial no encontrado en repo')
                log['xlsx_prov_trim_missing'] = True
                log['xlsx_prov_trim_missing_date'] = today

            if bytes_nac and bytes_dept:
                print('  Construyendo data.json...')
                import sys; sys.path.insert(0, 'scripts')
                from generar_empleo import build_empleo
                empleo = build_empleo(bytes_nac, bytes_dept, bytes_xlsx, bytes_prov_trim)
                with open(EMP_PATH, 'w', encoding='utf-8') as f:
                    json.dump(empleo, f, ensure_ascii=False, separators=(',', ':'))
                print(f'  ✓ data.json actualizado — último período: {empleo["meta"]["ultimo_sipa"]}')
                log['oede_signature'] = oede_sig
                log['oede_last_update'] = today
                updated = True
            else:
                print('  ✗ No se pudieron bajar todos los archivos')
        else:
            print(f'  Sin cambios en OEDE (última actualización: {log.get("oede_last_update", "nunca")})')
    except Exception as e:
        print(f'  ✗ Error en OEDE: {e}')

    log['last_check'] = today
    save_log(log)

    if updated:
        print(f'\n✓ Datos actualizados el {today}')
    else:
        print(f'\n— Sin cambios el {today}')

if __name__ == '__main__':
    main()
