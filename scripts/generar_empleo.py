"""
scripts/generar_empleo.py
Genera data.json desde el archivo SIPA mensual de trabajo registrado.
Llamado por actualizar.py cuando detecta datos nuevos.

Fuentes (todas del mismo XLSX):
  - A.2.1 → sectores nacionales con estacionalidad (en miles, puestos absolutos)
  - A.2.2 → sectores nacionales desestacionalizados (en miles)
  - A.5.1 → empleo por provincia con estacionalidad (en miles)
  - A.5.2 → empleo por provincia desestacionalizado (en miles)
  + CSV departamental → departamentos + sectores provinciales
"""

import re, json
from io import BytesIO
from datetime import datetime
from collections import defaultdict

# ── Mapeos ────────────────────────────────────────────────────────────────────
PROV_MAP = {
    'BUENOS AIRES':                        'Buenos Aires',
    'Cdad. Autónoma \nde Buenos Aires':    'C.A.B.A.',
    'CATAMARCA 3/':                        'Catamarca',
    'CHACO':                               'Chaco',
    'CHUBUT':                              'Chubut',
    'CÓRDOBA':                             'Córdoba',
    'CORRIENTES':                          'Corrientes',
    'ENTRE RÍOS':                          'Entre Ríos',
    'FORMOSA':                             'Formosa',
    'JUJUY':                               'Jujuy',
    'LA PAMPA':                            'La Pampa',
    'LA RIOJA':                            'La Rioja',
    'MENDOZA':                             'Mendoza',
    'MISIONES':                            'Misiones',
    'NEUQUÉN':                             'Neuquén',
    'RíO NEGRO':                           'Río Negro',
    'SALTA':                               'Salta',
    'SAN JUAN':                            'San Juan',
    'SAN LUIS':                            'San Luis',
    'SANTA CRUZ':                          'Santa Cruz',
    'SANTA FE':                            'Santa Fe',
    'SANTIAGO \nDEL ESTERO':              'Santiago del Estero',
    'TIERRA DEL FUEGO':                    'Tierra del Fuego',
    'TUCUMÁN':                             'Tucumán',
}

PROV_ID = {
    'Buenos Aires':'06','C.A.B.A.':'02','Catamarca':'10','Chaco':'16',
    'Chubut':'26','Córdoba':'14','Corrientes':'18','Entre Ríos':'30',
    'Formosa':'34','Jujuy':'38','La Pampa':'42','La Rioja':'46',
    'Mendoza':'50','Misiones':'54','Neuquén':'58','Río Negro':'62',
    'Salta':'66','San Juan':'70','San Luis':'74','Santa Cruz':'78',
    'Santa Fe':'82','Santiago del Estero':'86','Tierra del Fuego':'94',
    'Tucumán':'90',
}

# Mapeo ramas SIPA → 7 macrosectores (nombres exactos de A.2.1)
MACRO_MAP = {
    'Agro y pesca':           ['Agricultura, ganaderÍa, \ncaza y silvicultura', 'Pesca'],
    'Minas y petróleo':       ['Explotación de \nminas y canteras'],
    'Industria manufacturera':['Industrias manufactureras'],
    'Electricidad, gas y agua':['Suministro de electricidad, \ngas y agua'],
    'Construcción':           ['Construcción'],
    'Comercio':               ['Comercio y reparaciones'],
    'Servicios':              ['Hoteles y restaurantes',
                               'Transporte, almacenamiento\n y comunicación',
                               'Intermediación financiera',
                               'Actividades inmobiliarias, \nempresariales y de alquiler',
                               'Enseñanza',
                               'Servicios sociales \ny de salud',
                               'Servicios comunitarios, \nsociales y personales'],
}
RAMA_MACRO = {rama: macro for macro, ramas in MACRO_MAP.items() for rama in ramas}

MESES = {'ene':'01','feb':'02','mar':'03','abr':'04','may':'05','jun':'06',
         'jul':'07','ago':'08','sep':'09','oct':'10','nov':'11','dic':'12'}

def parse_periodo(s):
    s = str(s).strip().rstrip('*').strip()
    if isinstance(s, str) and re.match(r'[a-z]{3}-\d{2}$', s, re.I):
        mes, anio2 = s[:3].lower(), s[4:]
        mes_n = MESES.get(mes)
        if mes_n:
            anio = ('20' if int(anio2) < 50 else '19') + anio2
            return f'{anio}-{mes_n}'
    return None

def is_data_row(val):
    if val is None: return False
    if isinstance(val, datetime): return True
    s = str(val).strip()
    return bool(re.match(r'[a-z]{3}-\d{2}', s, re.I))

def parse_sheet(ws, n_cols=None):
    """Lee una hoja y devuelve {periodo: [v1, v2, ...]} y lista de columnas."""
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1]
    cols = []
    for i, v in enumerate(header):
        if v and i > 0:
            cols.append((i, str(v).strip()))
        if n_cols and len(cols) >= n_cols:
            break

    result = {}
    for row in rows[2:]:
        cell = row[0]
        if isinstance(cell, datetime):
            t = cell.strftime('%Y-%m')
        else:
            t = parse_periodo(cell)
        if not t:
            continue
        vals = []
        for i, _ in cols:
            v = row[i] if i < len(row) else None
            vals.append(round(v * 1000) if isinstance(v, (int, float)) else None)
        result[t] = vals
    return result, [c for _, c in cols]


def parse_sipa(bytes_xlsx):
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(bytes_xlsx), data_only=True, read_only=True)

    # A.2.1 — sectores nacionales con estacionalidad
    sec_orig, sec_names = parse_sheet(wb['A.2.1'], n_cols=14)  # excluir Sin especificar y Total
    # A.2.2 — sectores nacionales desestacionalizados
    sec_desa, _ = parse_sheet(wb['A.2.2'], n_cols=14)

    # A.1 — total nacional sector privado (orig col 1, desa col 2)
    ws_a1 = wb['A.1']
    rows_a1 = list(ws_a1.iter_rows(values_only=True))
    nac_a1_orig = {}
    nac_a1_desa = {}
    for row in rows_a1[3:]:  # skip headers
        cell = row[0]
        if isinstance(cell, datetime):
            t = cell.strftime('%Y-%m')
        else:
            t = parse_periodo(cell)
        if not t: continue
        if isinstance(row[1], (int, float)):
            nac_a1_orig[t] = round(row[1] * 1000)
        if isinstance(row[2], (int, float)):
            nac_a1_desa[t] = round(row[2] * 1000)

    # A.5.1 — provincias con estacionalidad
    prov_orig, prov_names = parse_sheet(wb['A.5.1'])
    # A.5.2 — provincias desestacionalizadas
    prov_desa, _ = parse_sheet(wb['A.5.2'])

    wb.close()

    # Organizar provincias
    prov_series = {}
    for j, col_raw in enumerate(prov_names):
        nombre = None
        for k, v in PROV_MAP.items():
            if col_raw.startswith(k[:8]) or k.startswith(col_raw[:8]):
                nombre = v
                break
        if not nombre:
            nombre = PROV_MAP.get(col_raw)
        if not nombre:
            continue
        prov_series[nombre] = {
            'orig': {t: vals[j] for t, vals in prov_orig.items() if vals[j] is not None},
            'desa': {t: vals[j] for t, vals in prov_desa.items() if vals[j] is not None},
        }

    # Organizar sectores (macro + detalle)
    # Detalle: cada rama fina
    detalle_orig = {}  # rama → {t: v}
    detalle_desa = {}
    for j, rama in enumerate(sec_names):
        if 'Sin especificar' in rama or 'Total' in rama:
            continue
        detalle_orig[rama] = {t: vals[j] for t, vals in sec_orig.items() if vals[j] is not None}
        detalle_desa[rama] = {t: vals[j] for t, vals in sec_desa.items() if vals[j] is not None}

    # Macro: sumar ramas al grupo
    macro_orig = defaultdict(lambda: defaultdict(int))
    macro_desa = defaultdict(lambda: defaultdict(int))
    for rama, grupo in RAMA_MACRO.items():
        for t, v in detalle_orig.get(rama, {}).items():
            macro_orig[grupo][t] += v
        for t, v in detalle_desa.get(rama, {}).items():
            macro_desa[grupo][t] += v

    return prov_series, dict(macro_orig), dict(macro_desa), detalle_orig, detalle_desa, nac_a1_orig, nac_a1_desa


def parse_departamental(bytes_csv, bytes_xlsx=None):
    import csv, io
    content = bytes_csv.decode('utf-8-sig', errors='replace')
    reader = csv.DictReader(io.StringIO(content), delimiter=';')

    SECTOR_MAP = {
        'Agricultura, ganaderia y pesca': 'Agro y pesca',
        'Explotacion de minas y canteras': 'Minas y petróleo',
        'Industria manufacturera': 'Industria',
        'Electricidad, gas y agua': 'Electricidad, gas y agua',
        'Construccion': 'Construcción',
        'Comercio': 'Comercio',
        'Servicios': 'Servicios',
    }

    deptos = defaultdict(lambda: {
        'label':'','provincia':'','id_prov':'',
        'serie':{},'serie_sec':{},'serie_xlsx':{},'sectores':defaultdict(dict)
    })
    prov_sec = defaultdict(lambda: defaultdict(dict))  # prov → sector → {t: v}

    PROV_NAME_MAP = {
        'CABA': 'C.A.B.A.',
        'BUENOS AIRES': 'Buenos Aires',
        '40 MUNICIPIOS GBA': 'Buenos Aires',  # se acumula en BA
        'RESTO DE PBA': 'Buenos Aires',        # se acumula en BA
        'CATAMARCA': 'Catamarca',
        'CHACO': 'Chaco',
        'CHUBUT': 'Chubut',
        'CORDOBA': 'Córdoba',
        'CORRIENTES': 'Corrientes',
        'ENTRE RIOS': 'Entre Ríos',
        'FORMOSA': 'Formosa',
        'JUJUY': 'Jujuy',
        'LA PAMPA': 'La Pampa',
        'LA RIOJA': 'La Rioja',
        'MENDOZA': 'Mendoza',
        'MISIONES': 'Misiones',
        'NEUQUEN': 'Neuquén',
        'RIO NEGRO': 'Río Negro',
        'SALTA': 'Salta',
        'SAN JUAN': 'San Juan',
        'SAN LUIS': 'San Luis',
        'SANTA CRUZ': 'Santa Cruz',
        'SANTA FE': 'Santa Fe',
        'SANTIAGO DEL ESTERO': 'Santiago del Estero',
        'TIERRA DEL FUEGO': 'Tierra del Fuego',
        'TUCUMAN': 'Tucumán',
    }

    for row in reader:
        try:
            codigo = str(row.get('id_depto', row.get('codigo_departamento',''))).strip().zfill(5)
            if not codigo or codigo == '00000': continue
            periodo_raw = row.get('Periodo', row.get('periodo','')).strip()
            t = None
            if re.match(r'\d{4}-\d{2}', periodo_raw): t = periodo_raw[:7]
            elif re.match(r'\d{6}', periodo_raw): t = periodo_raw[:4]+'-'+periodo_raw[4:6]
            if not t: continue

            nombre_depto = row.get('Departamento', row.get('nombre_departamento','')).strip()
            nombre_prov  = row.get('Provincia', row.get('nombre_provincia','')).strip()
            id_prov      = str(row.get('id_prov', row.get('codigo_provincia',''))).strip().zfill(2)
            sector_raw   = row.get('Sector', row.get('sector', row.get('letra_actividad',''))).strip()
            sector       = SECTOR_MAP.get(sector_raw, sector_raw)

            empleo_str = str(row.get('Empleo', row.get('empleo', row.get('puestos', row.get('trabajadores',''))))).strip()
            if not empleo_str: continue
            empleo = round(float(empleo_str.replace(',','.')))  # CSV en puestos reales

            d = deptos[codigo]
            d['label'] = f'{nombre_depto} — {nombre_prov}'
            d['provincia'] = nombre_prov
            d['id_prov'] = id_prov

            if not sector_raw or sector_raw in ('Total', 'Sin rama', '0'):
                pass  # ignorar Sin rama — usamos suma de sectores
            else:
                d['sectores'][sector][t] = empleo
                # Total = suma de sectores (puestos reales)
                d['serie_sec'][t] = d['serie_sec'].get(t, 0) + empleo
                if nombre_prov:
                    prov_norm = PROV_NAME_MAP.get(nombre_prov.upper(), nombre_prov)
                    prov_sec[prov_norm][sector][t] = \
                        prov_sec[prov_norm][sector].get(t, 0) + empleo
        except (ValueError, KeyError):
            continue

    # Si tenemos el XLSX, usarlo para el total (más preciso)
    if bytes_xlsx:
        import openpyxl
        from datetime import datetime as dt2
        wb = openpyxl.load_workbook(BytesIO(bytes_xlsx), data_only=True, read_only=True)
        for sheet in wb.sheetnames:
            if not sheet.startswith('T'): continue
            try: int(sheet[1:])
            except: continue
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2: continue
            header = rows[1]
            # Find date columns
            date_cols = [(i, v.strftime('%Y-%m')) for i, v in enumerate(header)
                        if isinstance(v, dt2)]
            if not date_cols: continue
            # Skip remuneraciones sheets (T7-T12 typically)
            title = str(rows[0][0] if rows[0] else '')
            if 'Remuner' in title: continue
            for row in rows[2:]:
                codigo = str(row[1]).strip().zfill(5) if row[1] else None
                if not codigo or codigo == '00000': continue
                nombre_depto = str(row[0]).strip() if row[0] else ''
                nombre_prov  = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                id_prov = codigo[:2]
                d = deptos[codigo]
                d['label'] = f'{nombre_depto} — {nombre_prov}'
                d['provincia'] = nombre_prov
                d['id_prov'] = id_prov
                # Store XLSX values separately (puestos reales, más precisos)
                for col_i, t in date_cols:
                    v = row[col_i] if col_i < len(row) else None
                    if isinstance(v, (int, float)) and v > 0:
                        d['serie_xlsx'][t] = round(v)
        wb.close()

    return deptos, dict(prov_sec)


def delta_pct(serie_dict, ini, fin):
    base = serie_dict.get(ini)
    ult  = serie_dict.get(fin) or serie_dict.get(
        max((t for t in serie_dict if t <= fin), default=''))
    if base and ult and base > 0:
        return ult - base, round((ult - base) / base * 100, 1)
    return None, None


def build_empleo(bytes_sipa, bytes_dept, bytes_xlsx=None):
    print('  Parseando SIPA (A.2.1/A.2.2/A.5.1/A.5.2)...')
    prov_series, macro_orig, macro_desa, detalle_orig, detalle_desa, nac_orig, nac_desa = parse_sipa(bytes_sipa)

    print('  Parseando CSV departamental...')
    deptos, prov_sec = parse_departamental(bytes_dept, bytes_xlsx)

    # Períodos
    sample = next(iter(prov_series.values()))
    all_t = sorted(sample['orig'].keys())
    ultimo_sipa = all_t[-1]

    all_t_dept = sorted({t for d in deptos.values() for t in d['serie']})
    ultimo_depto = all_t_dept[-1] if all_t_dept else ultimo_sipa

    print(f'  SIPA hasta: {ultimo_sipa} | Depto hasta: {ultimo_depto}')

    # Nacional viene de A.1 (ya parseado en parse_sipa)

    PRESIDENCIAS_CFG = {
        'Alberto Fernández': ('2019-11', '2023-11'),
        'Milei':             ('2023-11', ultimo_sipa),
    }

    presidencias_out = {}
    for presi, (ini, fin) in PRESIDENCIAS_CFG.items():
        fin_dept = min(fin, ultimo_depto)

        # ── Sectores nacionales ──
        def build_sectores(sec_dict, ini, fin):
            """Agrupa las 14 ramas del SIPA en 7 macros usando MACRO_MAP."""
            secs = []
            for macro, ramas in MACRO_MAP.items():
                delta_total = 0
                base_total = 0
                found = False
                for rama in ramas:
                    sd = sec_dict.get(rama, {})
                    b = sd.get(ini)
                    u = sd.get(fin) or sd.get(max((t for t in sd if t <= fin), default=''))
                    if b and u:
                        delta_total += u - b
                        base_total += b
                        found = True
                if found and base_total > 0:
                    pct = round(delta_total / base_total * 100, 1)
                    secs.append({'sector': macro, 'delta': delta_total, 'pct': pct})
            return sorted(secs, key=lambda x: x['delta'])

        def build_detalle(detalle, ini, fin):
            rows = []
            for rama, sd in detalle.items():
                d, p = delta_pct(sd, ini, fin)
                if d is not None:
                    rows.append({'sector': rama, 'delta': d, 'pct': p})
            return sorted(rows, key=lambda x: x['delta'])

        nac_delta, nac_pct = delta_pct(dict(nac_orig), ini, fin)
        nac_delta_d, nac_pct_d = delta_pct(dict(nac_desa), ini, fin)

        def slice_t(d, ini, fin):
            return [{'t':t,'v':v} for t,v in sorted(d.items()) if ini<=t<=fin]

        # ── Provincias ──
        provincias_out = {}
        for nombre, ps in prov_series.items():
            d, p     = delta_pct(ps['orig'], ini, fin)
            d_d, p_d = delta_pct(ps['desa'], ini, fin)

            # Sectores provinciales del departamental
            ps_sec = prov_sec.get(nombre, {})
            secs_prov = []
            for sector, sd in ps_sec.items():
                sd2, sp2 = delta_pct(sd, ini, fin_dept)
                if sd2 is not None:
                    secs_prov.append({'sector': sector, 'delta': sd2, 'pct': sp2})
            secs_prov.sort(key=lambda x: x['delta'])

            # Sub de Buenos Aires (GBA / Resto) si disponible
            sub = None
            if nombre == 'Buenos Aires':
                sub = {}
                for sub_key in ['gba', 'resto']:
                    sub_deptos = {
                        k: v for k, v in deptos.items()
                        if v['id_prov'] == PROV_ID[nombre] and
                        (sub_key == 'gba') == _is_gba(k)
                    }
                    sub_orig = defaultdict(int)
                    for dv in sub_deptos.values():
                        for t, v in dv['serie'].items():
                            sub_orig[t] += v
                    sd2, sp2 = delta_pct(dict(sub_orig), ini, fin_dept)
                    sub_secs = []
                    sub_sec_agg = defaultdict(lambda: defaultdict(int))
                    for dv in sub_deptos.values():
                        for sec, st in dv['sectores'].items():
                            for t, v in st.items():
                                sub_sec_agg[sec][t] += v
                    for sec, st in sub_sec_agg.items():
                        sd3, sp3 = delta_pct(dict(st), ini, fin_dept)
                        if sd3 is not None:
                            sub_secs.append({'sector': sec, 'delta': sd3, 'pct': sp3})
                    sub[sub_key] = {
                        'delta': sd2, 'pct': sp2,
                        'serie': slice_t(dict(sub_orig), ini, fin_dept),
                        'sectores': sorted(sub_secs, key=lambda x: x['delta']),
                    }

            prov_obj = {
                'id_prov':    PROV_ID.get(nombre, '00'),
                'base':       ps['orig'].get(ini),
                'ult':        ps['orig'].get(fin),
                'delta':      d,
                'pct':        p,
                'base_desa':  ps['desa'].get(ini),
                'ult_desa':   ps['desa'].get(fin),
                'delta_desa': d_d,
                'pct_desa':   p_d,
                'serie':      slice_t(ps['orig'], ini, fin),
                'serie_desa': slice_t(ps['desa'], ini, fin),
                'sectores':   secs_prov,
                'detalle':    [],
            }
            if sub:
                prov_obj['sub'] = sub
            provincias_out[nombre] = prov_obj

        # ── Departamentos ──
        depts_out = {}
        for codigo, dv in deptos.items():
            # Total: XLSX si disponible (más preciso), sino suma de sectores del CSV
            serie_total = dv['serie_xlsx'] if dv.get('serie_xlsx') else dv.get('serie_sec', {})
            if not serie_total: continue
            d2, p2 = delta_pct(serie_total, ini, fin_dept)
            secs = []
            for sec, st in dv['sectores'].items():
                sd2, sp2 = delta_pct(dict(st), ini, fin_dept)
                if sd2 is not None:
                    secs.append({'sector': sec, 'delta': sd2, 'pct': sp2})
            serie_dep = dv['serie_xlsx'] if dv.get('serie_xlsx') else dv.get('serie_sec', {})
            depts_out[codigo] = {
                'label':    dv['label'],
                'provincia': dv['provincia'],
                'id_prov':  dv['id_prov'],
                'delta':    d2,
                'pct':      p2,
                'serie':    slice_t(serie_dep, ini, fin_dept),
                'sectores': sorted(secs, key=lambda x: x['delta']),
            }

        nac_base = nac_orig.get(ini, 0)
        nac_ult  = nac_orig.get(fin, 0)

        presidencias_out[presi] = {
            'pais': {
                'base':        nac_base,
                'ult':         nac_ult,
                'delta':       nac_delta,
                'pct':         nac_pct,
                'base_desa':   nac_desa.get(ini),
                'ult_desa':    nac_desa.get(fin),
                'delta_desa':  nac_delta_d,
                'pct_desa':    nac_pct_d,
                'serie':       slice_t(dict(nac_orig), ini, fin),
                'serie_desa':  slice_t(dict(nac_desa), ini, fin),
                'sectores':    build_sectores(detalle_orig, ini, fin),
                'sectores_desa': build_sectores(detalle_desa, ini, fin),
                'detalle':     build_detalle(detalle_orig, ini, fin),
                'detalle_desa': build_detalle(detalle_desa, ini, fin),
            },
            'provincias':    provincias_out,
            'departamentos': depts_out,
        }

    return {
        'meta': {
            'ultimo_sipa':  ultimo_sipa,
            'ultimo_depto': ultimo_depto,
            'ultimo':       ultimo_sipa,
            'periodos': {p: list(cfg) for p, cfg in PRESIDENCIAS_CFG.items()},
            'fuente':       'OEDE / SIPA — Ministerio de Capital Humano',
            'actualizado':  datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'),
        },
        'presidencias': presidencias_out,
    }


# Codigos GBA (partidos del Gran Buenos Aires)
GBA_CODES = {
    '06028','06035','06091','06134','06245','06252','06260','06270',
    '06274','06277','06322','06329','06365','06371','06408','06410',
    '06412','06427','06434','06490','06515','06525','06539','06560',
    '06568','06638','06648','06658','06749','06756','06760','06778',
    '06805','06840','06861','06868','06882','06896','06897','06910',
    '06931',
}

def _is_gba(codigo):
    return codigo in GBA_CODES
